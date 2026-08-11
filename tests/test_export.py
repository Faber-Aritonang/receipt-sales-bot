"""Unit test ekspor Excel (app/export.py)."""
import io

import openpyxl

from app import export


def test_build_xlsx_valid(seed_data):
    data = export.build_xlsx()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Struk", "Item"]
    # header + 2 struk
    assert wb["Struk"].max_row - 1 == 2
    # header + 3 item
    assert wb["Item"].max_row - 1 == 3
    # header baris pertama tidak kosong
    assert wb["Struk"].cell(row=1, column=1).value


def test_build_xlsx_db_kosong(tmp_env):
    data = export.build_xlsx()
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Struk", "Item"]
    assert wb["Struk"].max_row - 1 == 0


def test_export_filename():
    name = export.export_filename()
    assert name.startswith("penjualan_")
    assert name.endswith(".xlsx")


def test_api_export_xlsx(api_client):
    r = api_client.get("/api/export/xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "attachment" in r.headers["content-disposition"]
    assert r.content.startswith(b"PK")  # signature file zip/xlsx
